from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
from openpyxl.utils.cell import column_index_from_string
import locale

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF8')

mes = datetime.now().strftime('%B').capitalize()
ano = datetime.now().strftime('%Y')
tamanhos_cabecalho = {'Preço Unitário': 5, 'Mão de Obra': 5, 'Composições Analíticas': 3, 'ABC de Serviços': 4, 'Resumido': 4, 'Orçamento Sintético': 4, 'Orçamento Analítico': 3, 'ABC de Insumos': 5}
colunas_titulo = {'Preço Unitário': 'C', 'Mão de Obra': 'D', 'Composições Analíticas': 'C', 'ABC de Serviços': 'C', 'Resumido': 'D', 'Orçamento Sintético': 'D', 'Orçamento Analítico': 'C', 'ABC de Insumos': 'C'}
banco_dados = {'Preço Unitário': 'E1:F2', 'Mão de Obra': 'E1:G2', 'Composições Analíticas': 'E1:G2', 'ABC de Serviços': 'D1:D2','Resumido': 'E1:E2', 'Orçamento Sintético': 'E1:F2', 'Orçamento Analítico': 'E1:F2', 'ABC de Insumos': 'D1:D2'}
nomes_relatorios = {'Preço Unitário': {'ordem': '09 - ', 'relatorio': ' - Composições com Preço Unitário'}, 'Mão de Obra': {'ordem': '08 - ', 'relatorio': ' - Sintético com Valor da Mão de Obra, Equipamento e Material'}, 'Composições Analíticas': {'ordem': '07 - ', 'relatorio': ' - Composições Analíticas com Detalhamento de LS e BDI'},'ABC de Serviços': {'ordem': '06 - ', 'relatorio': ' - Curva ABC de Serviços'}, 'ABC de Insumos': {'ordem': '05 - ', 'relatorio': ' - Curva ABC de Insumos'}, 'Orçamento Analítico': {'ordem': '04 - ', 'relatorio': ' - Orçamento Analítico'}, 'Resumido': {'ordem': '02 - ', 'relatorio': ' - Orçamento Resumido'}, 'Orçamento Sintético': {'ordem': '03 - ', 'relatorio': ' - Orçamento Sintético'}}
coluna_final_filtro = {'ABC de Insumos': 'O', 'ABC de Serviços': 'J'}

def formatar_relatorio_analitico_preco_unit(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
  
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  salvar_relatorio(planilha, file, empreendimento)
  
def formatar_relatorio_sintetico_mo_eq_mat(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
  
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  aumentar_largura_coluna(aba, ['O'], tamanho_cabecalho)
  
  salvar_relatorio(planilha, file, empreendimento)

def formatar_relatorio_analitico_somente_insumos(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
  
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  salvar_relatorio(planilha, file, empreendimento)

def formatar_relatorio_abc_servicos(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
  
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  aumentar_largura_coluna(aba, ['G','H'], tamanho_cabecalho)
  
  maxRow = aplicar_filtro(aba, file, tamanho_cabecalho)
  
  salvar_relatorio(planilha, file, empreendimento)

def formatar_relatorio_abc_insumos(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
    
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  aumentar_largura_coluna(aba, ['N'], tamanho_cabecalho)
  
  maxRow = aplicar_filtro(aba, file, tamanho_cabecalho)
  
  verifica_valor_zerado(aba, ['F', 'H'], tamanho_cabecalho, maxRow)
  
  excluir_coluna(aba, ['P', 'Q'])
  
  salvar_relatorio(planilha, file, empreendimento)

def formatar_relatorio_analitico(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
  
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  aumentar_largura_coluna(aba, ['H', 'J'], tamanho_cabecalho)
  
  salvar_relatorio(planilha, file, empreendimento)

def formatar_relatorio_sintetico(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
  
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  aumentar_largura_coluna(aba, ['I', 'J'], tamanho_cabecalho)
  
  salvar_relatorio(planilha, file, empreendimento)

def formatar_relatorio_resumido(file, base_dados):
  planilha = load_workbook(file)
  aba = planilha.active
  
  modifica_banco_dados(file, aba, base_dados)
  
  excluir_campo_assinatura(aba)  
  
  # Seleciona o tamanho do cabeçalho
  tamanho_cabecalho = verifica_tamanho_cabecalho(file)
  
  fixar_linhas_cabecalho(aba, tamanho_cabecalho)
  
  congelar_paineis(aba, tamanho_cabecalho)
  
  empreendimento = remover_nome_bloqueado(file, aba)
  
  aumentar_largura_coluna(aba, ['J', 'K'], tamanho_cabecalho)
  
  salvar_relatorio(planilha, file, empreendimento)

def excluir_coluna(aba, colunas):
  for col in colunas:
    aba.delete_cols(column_index_from_string(col), 2)

def aplicar_filtro(aba, file, tamanho_cabecalho):
  coluna = ''
  for key in nomes_relatorios:
    if file.find(key) != -1:
      coluna = str(coluna_final_filtro[key])
      
  
  maxRow = 0
  for cell in aba['A']:
    if cell.row > tamanho_cabecalho:
      if cell.value is not None:
          maxRow = cell.row
  
  aba.auto_filter.ref = f'A{tamanho_cabecalho}:{coluna}{maxRow}'
  
  return maxRow

def salvar_relatorio(planilha, file, empreendimento):
  ordem = ''
  for key in nomes_relatorios:
    if file.find(key) != -1:
      ordem = nomes_relatorios[key]['ordem']
      relatorio = nomes_relatorios[key]['relatorio']
  
  titulo = ordem + empreendimento + relatorio + '.xlsx'
  
  # planilha.save(titulo)
  # planilha = load_workbook(titulo)
  aba = planilha.active
  remover_exibicao_linhas_grade(aba)
  planilha.save(titulo)
  
  print('-> Concluído: ' + ordem + empreendimento + relatorio)

def aumentar_largura_coluna(aba, colunas_aumentar, tamanho_cabecalho):
  for col in colunas_aumentar:
    # Identifica maior largura da coluna
    max_length = 0
    n = 0
    for cel in aba[col]:
      if (n >= 10): break
      if cel.row > tamanho_cabecalho - 1:
        if str(cel.value).find(".") != -1 and isinstance(cel.value, (int, float)) and len(str(cel.value)) > 4:
          if (len(str(cel.value).split(".")[1]) > 2):
            length = len(str(cel.value).split(".")[0] + "." + str(cel.value).split(".")[1][0:1])
          else:
            length = len(str(cel.value))
          if length > max_length:
            max_length = length
        else:
          if len(str(cel.value)) > max_length:
            max_length = len(str(cel.value))
      n += 1
    # Atribui largura para a coluna
    aba.column_dimensions[col].width = max_length + 2

def remover_nome_bloqueado(file, aba):
  for key in colunas_titulo:
    if file.find(key) != -1:
      coluna_titulo = colunas_titulo[key]
  
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' - "BLOQUEADO"', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' - "Bloqueado"', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' - "bloqueado"', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' - BLOQUEADO', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' - Bloqueado', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' - bloqueado', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' BLOQUEADO', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' bloqueado', '')
  aba[f'{coluna_titulo}2'].value = aba[f'{coluna_titulo}2'].value.replace(' Bloqueado', '')
  return aba[f'{coluna_titulo}2'].value

def congelar_paineis(aba, tamanho_cabecalho):      
  aba.freeze_panes = f'A{tamanho_cabecalho + 1}'

def verifica_tamanho_cabecalho(file):
  for key in tamanhos_cabecalho:
    if file.find(key) != -1:
      return tamanhos_cabecalho[key]

def fixar_linhas_cabecalho(aba, tamanho_cabecalho):
  aba.print_title_rows = f'1:{tamanho_cabecalho}'

def colocar_logo(aba):
  img = Image('logo_BC.png')
  img.width *= 0.7
  img.height *= 0.7
  img.anchor = 'A1'
  aba.add_image(img)

def excluir_campo_assinatura(aba):
  # Identifica útilma linha
  max_row = aba.max_row
  # Pega o valor da última célula da coluna A
  celA_Max = aba[f"A{max_row}"].value
  
  # Verifica se a última célula da coluna A é 'None'
  if celA_Max is not None:
    if (celA_Max.find("_") != -1):
      # Deleta últimas 2 linhas
      aba.delete_rows(max_row - 1,2)

def remover_exibicao_linhas_grade(aba):
  # Remover linhas de grade
  aba.sheet_view.showGridLines = False

def modifica_banco_dados(file, aba, base_dados):
  # Desmesclar células
  for key in banco_dados:
    if file.find(key) != -1:
      mergedcells =[]
      range = ''
      
      for group in aba.merged_cells.ranges:
        mergedcells.append(group)
      
      for key in banco_dados:
        if file.find(key) != -1:
          range = str(banco_dados[key])
          row_ini = aba[str(range[0:2])].row
          col_ini = aba[str(range[0:2])].column
          row_fim = aba[str(range[3:5])].row
          col_fim = aba[str(range[3:5])].column
      
      for group in mergedcells:
        min_col, min_row, max_col, max_row = group.bounds
        
        if (min_row >= row_ini and max_row <= row_fim and min_col >= col_ini and max_col <= col_fim):
          aba.unmerge_cells(str(group))
      # Mesclar célula "Bancos" com a célula abaixo
      aba.merge_cells(range)
      # Modifica texto do banco de dados
      aba[range[0:2]].value = "Banco de Dados:\n" + base_dados['mes'] + " / " + base_dados['ano']

# def copiar_arquivo(fonte, destino, tamanho_cabecalho):
#   if os.path.exists(destino):
#     os.remove(destino)
#   shutil.copy2(fonte, destino)

def verifica_valor_zerado(aba, colunas, tamanho_cabecalho, maxRow):
  for col in colunas:
    for cel in aba[col]:
      if cel.row > maxRow:
        break
      if cel.row > tamanho_cabecalho - 1:
        if cel.value is not None:
          if (cel.value == '0,00' or cel.value == '0,01'):
            print('-> Valor zerado encontrado: '+ cel.value +'\nCélula ' + col + str(cel.row))
            exit()

